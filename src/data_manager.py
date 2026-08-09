

import os
import sys
from datetime import datetime

import pandas as pd
from openpyxl import Workbook

COLUMNS = ["Fecha", "Categoria", "Descripcion", "Monto"]
SHEETS = ["Ingresos", "Egresos"]
CATEGORIAS_SHEET = "Categorias"

if getattr(sys, "frozen", False):
.
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

DEFAULT_PATH = os.path.join(BASE_DIR, "data", "planilla.xlsx")


def _filtrar_df(df: pd.DataFrame, fecha_inicio: str = None, fecha_fin: str = None, categoria: str = None) -> pd.DataFrame:
    """Filtra un DataFrame de movimientos por rango de fechas y/o categoría."""
    if df.empty:
        return df

    df = df.copy()
    df["Fecha"] = pd.to_datetime(df["Fecha"])

    if fecha_inicio:
        df = df[df["Fecha"] >= pd.to_datetime(fecha_inicio)]
    if fecha_fin:
        df = df[df["Fecha"] <= pd.to_datetime(fecha_fin)]
    if categoria and categoria != "Todas":
        df = df[df["Categoria"] == categoria]

    return df


def leer_movimientos_filtrados(
    fecha_inicio: str = None,
    fecha_fin: str = None,
    categoria: str = None,
    ruta: str = DEFAULT_PATH,
) -> dict:
    """Igual que leer_movimientos pero aplicando filtros de fecha/categoría."""
    hojas = leer_movimientos(ruta)
    return {
        nombre: _filtrar_df(df, fecha_inicio, fecha_fin, categoria)
        for nombre, df in hojas.items()
        if nombre in SHEETS
    }


def crear_planilla_si_no_existe(ruta: str = DEFAULT_PATH) -> None:
    """Crea el archivo Excel con las hojas Ingresos/Egresos si no existe."""
    if os.path.exists(ruta):
        return

    os.makedirs(os.path.dirname(ruta), exist_ok=True)

    wb = Workbook()

    primera_hoja = wb.active
    primera_hoja.title = SHEETS[0]
    primera_hoja.append(COLUMNS)

    for nombre_hoja in SHEETS[1:]:
        hoja = wb.create_sheet(title=nombre_hoja)
        hoja.append(COLUMNS)

    hoja_categorias = wb.create_sheet(title=CATEGORIAS_SHEET)
    hoja_categorias.append(["Categoria"])

    wb.save(ruta)


def registrar_movimiento(
    tipo: str,
    fecha: str,
    categoria: str,
    descripcion: str,
    monto: float,
    ruta: str = DEFAULT_PATH,
) -> None:
    """
    Agrega una fila nueva a la hoja correspondiente ("Ingresos" o "Egresos").

    tipo: "ingreso" o "egreso" (no distingue mayúsculas/minúsculas)
    fecha: string en formato YYYY-MM-DD (si viene vacío, se usa hoy)
    """
    crear_planilla_si_no_existe(ruta)

    hoja_destino = "Ingresos" if tipo.lower().startswith("i") else "Egresos"

    if not fecha:
        fecha = datetime.now().strftime("%Y-%m-%d")


    hojas = pd.read_excel(ruta, sheet_name=None)  

    nueva_fila = pd.DataFrame(
        [[fecha, categoria, descripcion, monto]], columns=COLUMNS
    )
    hojas[hoja_destino] = pd.concat(
        [hojas[hoja_destino], nueva_fila], ignore_index=True
    )

    with pd.ExcelWriter(ruta, engine="openpyxl") as writer:
        for nombre_hoja, df in hojas.items():
            df.to_excel(writer, sheet_name=nombre_hoja, index=False)


def leer_movimientos(ruta: str = DEFAULT_PATH) -> dict:
    """Devuelve un diccionario {"Ingresos": df, "Egresos": df}."""
    crear_planilla_si_no_existe(ruta)
    return pd.read_excel(ruta, sheet_name=None)


def resumen_por_categoria(
    tipo: str,
    fecha_inicio: str = None,
    fecha_fin: str = None,
    ruta: str = DEFAULT_PATH,
) -> pd.Series:
    """Suma de montos agrupados por categoría, para Ingresos o Egresos."""
    hojas = leer_movimientos(ruta)
    hoja = "Ingresos" if tipo.lower().startswith("i") else "Egresos"
    df = _filtrar_df(hojas[hoja], fecha_inicio, fecha_fin)
    if df.empty:
        return pd.Series(dtype=float)
    return df.groupby("Categoria")["Monto"].sum()


def resumen_mensual(
    fecha_inicio: str = None,
    fecha_fin: str = None,
    categoria: str = None,
    ruta: str = DEFAULT_PATH,
) -> pd.DataFrame:
    """
    Devuelve un DataFrame con columnas [Mes, Ingresos, Egresos]
    sumando los montos de cada hoja agrupados por mes (YYYY-MM),
    aplicando filtros opcionales de fecha y categoría.
    """
    hojas = leer_movimientos(ruta)
    resultado = {}

    for nombre_hoja in SHEETS:
        df = _filtrar_df(hojas[nombre_hoja], fecha_inicio, fecha_fin, categoria)
        if df.empty:
            continue
        df["Mes"] = df["Fecha"].dt.strftime("%Y-%m")
        agrupado = df.groupby("Mes")["Monto"].sum()
        resultado[nombre_hoja] = agrupado

    df_final = pd.DataFrame(resultado).fillna(0)
    df_final = df_final.sort_index()
    return df_final


def obtener_categorias(ruta: str = DEFAULT_PATH) -> list:
    """Devuelve la lista de categorías guardadas (ordenada alfabéticamente)."""
    crear_planilla_si_no_existe(ruta)
    df = pd.read_excel(ruta, sheet_name=CATEGORIAS_SHEET)
    return sorted(df["Categoria"].dropna().astype(str).unique().tolist())


def agregar_categoria(nombre: str, ruta: str = DEFAULT_PATH) -> None:
    """Agrega una categoría nueva a la lista guardada (si no existe ya)."""
    crear_planilla_si_no_existe(ruta)
    nombre = nombre.strip()
    if not nombre or nombre in obtener_categorias(ruta):
        return

    hojas = pd.read_excel(ruta, sheet_name=None)
    nueva_fila = pd.DataFrame([[nombre]], columns=["Categoria"])
    hojas[CATEGORIAS_SHEET] = pd.concat(
        [hojas[CATEGORIAS_SHEET], nueva_fila], ignore_index=True
    )

    with pd.ExcelWriter(ruta, engine="openpyxl") as writer:
        for nombre_hoja, df in hojas.items():
            df.to_excel(writer, sheet_name=nombre_hoja, index=False)
